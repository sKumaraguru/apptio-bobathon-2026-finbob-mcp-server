"""Excel file processing for assessment reports."""

import openpyxl
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
import io

from ..models.internal import SheetInfo


class ExcelProcessor:
    """
    Process Excel files to extract sheet data.

    Handles both formatted sheets (Executive Summary) and data sheets.
    """

    def __init__(self, file_path: str):
        """
        Initialize Excel processor.

        Args:
            file_path: Path to Excel file
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self._load_workbook()

    def _load_workbook(self) -> None:
        """Load Excel workbook."""
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path, data_only=True, read_only=True  # Get calculated values, not formulas
            )
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")

    def get_sheet_names(self) -> List[str]:
        """
        Get list of all sheet names in workbook.

        Returns:
            List of sheet names
        """
        if not self.workbook:
            return []
        return self.workbook.sheetnames

    def get_sheet_info(self, sheet_name: str) -> SheetInfo:
        """
        Get information about a sheet.

        Args:
            sheet_name: Name of sheet

        Returns:
            SheetInfo object

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.get_sheet_names():
            raise ValueError(f"Sheet '{sheet_name}' not found")

        sheet = self.workbook[sheet_name]

        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Check if empty
        is_empty = max_row <= 1 or max_col == 0

        # Check if has headers (first row has values)
        has_headers = False
        if not is_empty and max_row > 0:
            first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            has_headers = any(cell is not None and str(cell).strip() for cell in first_row)

        return SheetInfo(
            name=sheet_name, row_count=max_row, column_count=max_col, has_headers=has_headers, is_empty=is_empty
        )

    def sheet_to_csv(self, sheet_name: str, include_headers: bool = True) -> str:
        """
        Convert sheet to CSV format.

        Args:
            sheet_name: Name of sheet
            include_headers: Whether to include headers

        Returns:
            CSV string

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.get_sheet_names():
            raise ValueError(f"Sheet '{sheet_name}' not found")

        # Use pandas for efficient CSV conversion
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0 if include_headers else None)

        # Convert to CSV
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, header=include_headers)
        return csv_buffer.getvalue()

    def sheet_to_json(self, sheet_name: str, include_headers: bool = True) -> List[Dict[str, Any]]:
        """
        Convert sheet to JSON format (list of dictionaries).

        Args:
            sheet_name: Name of sheet
            include_headers: Whether to use first row as headers

        Returns:
            List of row dictionaries

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.get_sheet_names():
            raise ValueError(f"Sheet '{sheet_name}' not found")

        # Use pandas for efficient conversion
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0 if include_headers else None)

        # Convert to list of dictionaries
        return df.to_dict("records")

    def get_sheet_data_range(self, sheet_name: str, start_row: int, end_row: int, include_headers: bool = True) -> str:
        """
        Get a range of rows from a sheet as CSV.

        Args:
            sheet_name: Name of sheet
            start_row: Starting row (1-based)
            end_row: Ending row (1-based, inclusive)
            include_headers: Whether to include headers

        Returns:
            CSV string for the range

        Raises:
            ValueError: If sheet doesn't exist or range is invalid
        """
        if sheet_name not in self.get_sheet_names():
            raise ValueError(f"Sheet '{sheet_name}' not found")

        sheet_info = self.get_sheet_info(sheet_name)

        if start_row < 1 or end_row > sheet_info.row_count:
            raise ValueError(f"Invalid row range: {start_row}-{end_row}")

        # Read full sheet
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0 if sheet_info.has_headers else None)

        # Adjust for 0-based indexing
        if sheet_info.has_headers:
            # If has headers, row 1 is headers, data starts at row 2
            data_start = start_row - 2  # -1 for 0-based, -1 for header row
            data_end = end_row - 1
        else:
            data_start = start_row - 1
            data_end = end_row

        # Slice dataframe
        df_slice = df.iloc[max(0, data_start) : data_end]

        # Convert to CSV
        csv_buffer = io.StringIO()
        df_slice.to_csv(csv_buffer, index=False, header=include_headers)
        return csv_buffer.getvalue()

    def extract_key_value_pairs(self, sheet_name: str, key_column: int = 0, value_column: int = 1) -> Dict[str, Any]:
        """
        Extract key-value pairs from a sheet (e.g., Analyzed Facts).

        Args:
            sheet_name: Name of sheet
            key_column: Column index for keys (0-based)
            value_column: Column index for values (0-based)

        Returns:
            Dictionary of key-value pairs

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.get_sheet_names():
            raise ValueError(f"Sheet '{sheet_name}' not found")

        df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)  # No headers for key-value sheets

        # Extract key-value pairs
        result = {}
        for _, row in df.iterrows():
            if len(row) > max(key_column, value_column):
                key = row[key_column]
                value = row[value_column]

                if pd.notna(key) and pd.notna(value):
                    # Clean key
                    key_str = str(key).strip()
                    if key_str:
                        result[key_str] = value

        return result

    def close(self) -> None:
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()

    def __del__(self):
        """Cleanup on deletion."""
        self.close()


# Made with Bob
